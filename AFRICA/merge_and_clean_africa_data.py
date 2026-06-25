import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from datetime import datetime

def get_all_excel_files(root_folder):
    """Recursively find all Excel files in the folder and subfolders, excluding merged output"""
    excel_files = []
    for root, dirs, files in os.walk(root_folder):
        for file in files:
            # Skip the merged output files and PDF
            if (file.endswith(('.xlsx', '.xls')) and 
                not file.startswith('~$') and
                'Africa_Startups_Merged_Cleaned' not in file):
                excel_files.append(os.path.join(root, file))
    return excel_files

def read_excel_sheets(file_path):
    """Read all sheets from an Excel file and combine them"""
    try:
        xls = pd.ExcelFile(file_path)
        dfs = []
        for sheet in xls.sheet_names:
            if sheet.lower() != 'pdf':  # Skip PDF sheets if any
                df = pd.read_excel(file_path, sheet_name=sheet)
                df['_source_file'] = os.path.basename(file_path)
                df['_source_sheet'] = sheet
                dfs.append(df)
        if dfs:
            return pd.concat(dfs, ignore_index=True)
        return None
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None

def clean_data(df):
    """Perform data cleaning operations"""
    # Remove completely empty rows
    df = df.dropna(how='all')
    
    # Strip whitespace from all string columns
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str).str.strip()
    
    # Remove rows where all values are 'nan' or empty
    df = df.replace('nan', pd.NA)
    df = df.replace('', pd.NA)
    df = df.dropna(how='all')
    
    # Remove rows where the first non-metadata column is empty (likely header rows or empty rows)
    # Find the first non-metadata column
    first_data_col = None
    for col in df.columns:
        if col not in ['_source_file', '_source_sheet']:
            first_data_col = col
            break
    
    if first_data_col:
        df = df[df[first_data_col].notna()]
    
    # Standardize common column values (case insensitive for common fields)
    for col in df.columns:
        if df[col].dtype == 'object':
            # Try to detect and standardize Yes/No columns
            if df[col].astype(str).str.lower().isin(['yes', 'no']).sum() > df.shape[0] * 0.5:
                df[col] = df[col].astype(str).str.capitalize()
    
    return df.reset_index(drop=True)

def find_duplicates(df):
    """Identify and separate duplicate records"""
    # Use all columns except the metadata columns for duplicate detection
    metadata_cols = ['_source_file', '_source_sheet']
    compare_cols = [col for col in df.columns if col not in metadata_cols]
    
    # Mark duplicates (keeping first occurrence as False)
    df['_is_duplicate'] = df[compare_cols].duplicated(keep='first')
    df['_first_occurrence'] = ''
    
    # For duplicates, record location of first unique occurrence
    # Create a helper column with row numbers for non-duplicates
    df_clean = df[~df['_is_duplicate']].copy()
    
    # Build a mapping of compare_cols to first occurrence location
    first_occurrence_map = {}
    for idx, row in df_clean.iterrows():
        key = tuple(row[compare_cols].values)
        first_occurrence_map[key] = f"{row['_source_file']} - Sheet: {row['_source_sheet']} (Row {idx + 2})"
    
    # Apply the mapping to duplicates
    for idx in df[df['_is_duplicate']].index:
        key = tuple(df.loc[idx, compare_cols].values)
        if key in first_occurrence_map:
            df.loc[idx, '_first_occurrence'] = first_occurrence_map[key]
    
    # Separate clean data from duplicates
    clean_data = df[~df['_is_duplicate']].copy()
    duplicates = df[df['_is_duplicate']].copy()
    
    # Remove all tracking columns from clean_data (keep only original columns)
    tracking_cols = ['_is_duplicate', '_source_file', '_source_sheet']
    clean_data = clean_data.drop(columns=tracking_cols)
    
    # Remove _is_duplicate and source columns from duplicates (keep only _first_occurrence)
    duplicates = duplicates.drop(columns=['_is_duplicate', '_source_file', '_source_sheet'])
    
    return clean_data, duplicates

def style_excel_sheets(writer, clean_data, duplicates):
    """Apply styling to the Excel sheets"""
    workbook = writer.book
    
    # Style for Clean Data sheet
    clean_sheet = workbook['Clean_Data']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in clean_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths for Clean Data
    for column in clean_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        clean_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Style for Duplicates sheet
    if len(duplicates) > 0:
        dup_sheet = workbook['Duplicates']
        for cell in dup_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Highlight duplicate rows
        dup_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        for row in dup_sheet.iter_rows(min_row=2, max_row=len(duplicates) + 1):
            for cell in row:
                cell.fill = dup_fill
        
        # Auto-adjust column widths for Duplicates
        for column in dup_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            dup_sheet.column_dimensions[column_letter].width = adjusted_width

def main():
    # Define the AFRICA folder path
    africa_folder = r"c:\Users\Nacu\Desktop\THESIS 1\AFRICA"
    
    print(f"Starting to merge and clean files from: {africa_folder}")
    print("=" * 80)
    
    # Get all Excel files
    excel_files = get_all_excel_files(africa_folder)
    print(f"\nFound {len(excel_files)} Excel files:")
    for file in excel_files:
        print(f"  - {os.path.relpath(file, africa_folder)}")
    
    # Read and combine all files
    print("\nReading and combining files...")
    all_data = []
    
    for file in excel_files:
        print(f"  Processing: {os.path.relpath(file, africa_folder)}")
        df = read_excel_sheets(file)
        if df is not None and len(df) > 0:
            all_data.append(df)
            print(f"    ✓ Read {len(df)} rows")
    
    if not all_data:
        print("\nNo data found in Excel files!")
        return
    
    # Combine all dataframes
    combined_df = pd.concat(all_data, ignore_index=True)
    print(f"\nTotal rows before cleaning: {len(combined_df)}")
    
    # Clean data
    print("Cleaning data...")
    cleaned_df = clean_data(combined_df)
    print(f"Total rows after cleaning: {len(cleaned_df)}")
    
    # Find duplicates
    print("Identifying duplicates...")
    clean_final, duplicates_df = find_duplicates(cleaned_df)
    
    print(f"\nFinal Statistics:")
    print(f"  - Unique records: {len(clean_final)}")
    print(f"  - Duplicate records: {len(duplicates_df)}")
    print(f"  - Total records: {len(clean_final) + len(duplicates_df)}")
    
    # Create output filename
    output_filename = f"Africa_Startups_Merged_Cleaned_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(africa_folder, output_filename)
    
    # Write to Excel with multiple sheets
    print(f"\nWriting to Excel file: {output_filename}")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        clean_final.to_excel(writer, sheet_name='Clean_Data', index=False)
        if len(duplicates_df) > 0:
            duplicates_df.to_excel(writer, sheet_name='Duplicates', index=False)
        
        # Apply styling
        style_excel_sheets(writer, clean_final, duplicates_df)
    
    print(f"✓ File saved successfully!")
    print(f"\nOutput file location:")
    print(f"  {output_path}")
    print("\nDone!")

if __name__ == "__main__":
    main()
